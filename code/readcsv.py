""" Script to read and manipulate the csv data files """

# Imports
from collections import OrderedDict
#from termcolor import colored
import itertools
import string
import math
#import xlwt
import openpyxl
import csv
import os


# Definitions
def read_csvfile(filepath):
    with open(filepath, 'rb') as f:
        reader_data = csv.reader(f, delimiter=',')
        data = list(reader_data)
        col_names = data[0]
        data = data[1:]
        data = data[6:] + data[:6]

        # Generate letters for columns
        l1 = list(string.ascii_uppercase)
        l2 = map(lambda x: ''.join(x), itertools.product(l1, repeat=2))
        l = l1 + l2

        # Map column Name, Column Letter and Serial Number
        for counter, (col, name) in enumerate(zip(l, col_names)):
            col_map[col] = counter
            col_name_map[col] = name
    return data


def create_data_bins(data):
    data_bins['365day'] = data
    data_bins['24hr'] = OrderedDict()
    counter = 1
    for i in xrange(0, 8759, 24): # 8760 - 1
        day = 'day_'+str(counter)
        data_bins['24hr'][day] = {}

        # Generate 1hr bins
        _1hr_bins = data_bins['24hr'][day]['1hr'] = data[i:i+24]

        # Generate 3hr bins.
        _3hr_bins = data_bins['24hr'][day]['3hr'] = []
        for j in xrange(0, len(_1hr_bins), 3):
            _3hr_bins.append(_1hr_bins[j:j+3])

        counter += 1

    data_bins['summer'] = data[3216:3960]   # Data division for 'Summer' Bin
    data_bins['winter'] = data[8015:8759]   # Data division for 'Winter' Bin


# Considering 24hr day as 7:00 AM to 6:00 AM
groups_info = {
    'Bed-Room': {'part-1' : {'3hr': [5,6,7]}},

    'Living-Room': {'part-1': {'3hr':[0]},
                    'part-2': {'3hr':[3,4]}},

    'Kitchen': {'part-1': {'3hr': [0]},
                'part-2': {'1hr': [6,5,4]}, # 11,12,13 -> 6,5,4(indices)
                'part-3': {'3hr': [4]}},

    'Half-Day-Blocks': {'part-1': {'3hr': [0,1,2,3]},
                    'part-2': {'3hr': [4,5,6,7]}}
}

def create_data_groups():
    group_data = data_bins['groups'] = {}

    for group, info in groups_info.items():
        tmp = []
        for item in info.values():
            for key, values in item.items():
                for value in values:
                    for i in xrange(1, 365+1):
                        if key == '3hr':
                            tmp.extend(data_bins['24hr']['day_'+str(i)]['3hr'][value])
                        elif key == '1hr':
                            tmp.append(data_bins['24hr']['day_'+str(i)]['1hr'][value])

        group_data[group] = tmp


# Re-check the Data
def recheck_data():
    #print colored("\n\n---------- Re-Checking the Input ----------", "green")
    print "\nKeys in Data Bin: " + ",".join(data_bins.keys())
    print "\n365day Bin Size: " + str(len(data_bins['365day']))
    print "\n24hr Bin Size: " + str(len(data_bins['24hr']))
    print "\nSummer: " + str(len(data_bins['summer']))
    print "\nWinter: " + str(len(data_bins['winter']))

    print "\n1hr Sub-Bin Size: " + str(len(data_bins['24hr']['day_1']['1hr']))
    print "\n3hr Sub-Bin Size: " + str(len(data_bins['24hr']['day_1']['3hr']))

    print "\nDifferent Groups: " + ",".join(data_bins['groups'].keys())
    print "\nBed-Room Group Size: " + str(len(data_bins['groups']['Bed-Room']))
    print "\nBed-Room Data Item Size: " + str(len(data_bins['groups']['Bed-Room'][0]))
    print "\nLiving-Room Group Size: " + str(len(data_bins['groups']['Living-Room']))
    print "\nLiving-Room Data Item Size: " + str(len(data_bins['groups']['Living-Room'][0]))
    print "\nKitchen Group Size: " + str(len(data_bins['groups']['Kitchen']))
    print "\nKitchen Data Item Size: " + str(len(data_bins['groups']['Kitchen'][0]))
    print "\nHalf-Day-Blocks Group Size: " + str(len(data_bins['groups']['Half-Day-Blocks']))
    print "\nHalf-Day-Blocks Data Items Size: " + str(len(data_bins['groups']['Half-Day-Blocks'][0]))


def create_excel_workbook():
    wb = openpyxl.Workbook()
    wb.save("Output.xlsx")

# Task Definitions
def do_task1():
    output_bins
    output_bins['365day'] = {}
    days = output_bins['365day']['days'] = OrderedDict()
    for i in xrange(0, 365):
        day = days['day_'+str(i+1)] = {}
        tmp = []
        rows_list = data_bins['24hr']['day_'+str(i+1)]['1hr']
        tmp.append(get_max('B', rows_list))
        tmp.append(get_min('B', rows_list))
        tmp.append(get_mean('B', rows_list))
        tmp.append(get_range('B', rows_list))
        day['B'] = tmp
    #print output_bins['365day']['days']['day_1']['B']
    return output_bins

def prepare_task1():
    sheet = "Task-1"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.create_sheet()
    ws.title = sheet
    ws.merge_cells("A1:E1")
    ws['A1'] = "Outside Drybulb Temperature"
    col_names = ['Day', 'Maximum', 'Minimum', 'Mean', 'Range']
    ws.append(col_names)
    ws.freeze_panes = 'A3'
    wb.save("Output.xlsx")

def export_task1(filename):
    sheet = "Task-1"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.get_sheet_by_name(sheet)

    for i, (key, value) in enumerate(output_bins['365day']['days'].items()):
        data = value['B']
        _data = [filename+'_day'+str(i+1), data[0], data[1], data[2], data[3]]
        ws.append(_data)
    wb.save("Output.xlsx")


wall_info = {
    'wall_1': ['C', 'D', 'G', 'H'],
    'wall_2': ['I', 'J', 'M', 'N'],
    'wall_3': ['O', 'P', 'S', 'T'],
    'wall_4': ['U', 'V', 'Y', 'Z'],
    'wall_5': ['AA', 'AB', 'AE', 'AF'],
    'wall_6': ['AG', 'AH', 'AK', 'AL'],
    'wall_7': ['AM', 'AN', 'AQ', 'AR'],
    'wall_8': ['AS', 'AT', 'AW', 'AX'],
    'wall_9': ['AY', 'AZ', 'BC', 'BD'],
    'wall_10': ['BE', 'BF', 'BI', 'BJ'],
    'wall_11': ['BK', 'BL', 'BO', 'BP'],
    'wall_12': ['BQ', 'BR', 'BU', 'BV']
}

def do_task2():
    block_s = output_bins['summer'] = {}
    block_w = output_bins['winter'] = {}
    days_s = block_s['days'] = OrderedDict()
    days_w = block_w['days'] = OrderedDict()

    counter = 1
    for i in xrange(0, 744, 24):
        day_s = days_s['day_'+str(counter)] = {}
        day_w = days_w['day_'+str(counter)] = {}
        walls_s = day_s['walls'] = OrderedDict()
        walls_w = day_w['walls'] = OrderedDict()
        for name, info in wall_info.items():
            wall_s = walls_s[name] = []
            wall_w = walls_w[name] = []
            for col in info:
                tmp_s = []
                tmp_w = []

                # Calculations for 'Summer' bin
                rows_list_s = data_bins['summer'][i:i+24]
                tmp_s.append(get_max(col, rows_list_s))
                tmp_s.append(get_min(col, rows_list_s))
                tmp_s.append(get_mean(col, rows_list_s))

                # Calculations for 'Winter' bin
                rows_list_w = data_bins['winter'][i:i+24]
                tmp_w.append(get_max(col, rows_list_w))
                tmp_w.append(get_min(col, rows_list_w))
                tmp_w.append(get_mean(col, rows_list_w))

                wall_s.append(tmp_s)
                wall_w.append(tmp_w)
        counter += 1
    return output_bins

def prepare_task2():
    sheet = "Task-2"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.create_sheet()
    ws.title = sheet
    col_names = ['Day'] + ['Maximum', 'Minimum', 'Mean'] * 4 * 12
    ws.append(col_names)
    ws.freeze_panes = 'A2'
    wb.save("Output.xlsx")

def export_task2(filename):
    sheet = "Task-2"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.get_sheet_by_name(sheet)

    for i in xrange(0, 31):
        data_s = output_bins['summer']['days']['day_'+str(i+1)]['walls']

        _data_s = [filename+'_day'+str(i+1)+'_summer']
        for value in data_s.values():
            for x in value:
                _data_s.extend(x)

        ws.append(_data_s)

    for i in xrange(0, 31):
        data_w = output_bins['winter']['days']['day_'+str(i+1)]['walls']

        _data_w = [filename+'_day'+str(i+1)+'_winter']
        for value in data_w.values():
            for x in value:
                _data_w.extend(x)

        ws.append(_data_w)
    wb.save("Output.xlsx")


Room_info = {
    'Room-1': ['CJ', 'DP'],
    'Room-2': ['CK', 'DQ'],
    'Room-3': ['CL', 'DR'],
    'Room-4': ['CM', 'DS'],
    'Room-5': ['CN', 'DU'],
    'Room-6': ['CU', 'DV'],
    'Room-7': ['CP', 'DW'],
    'Room-8': ['CQ', 'DX'],
    'Room-9': ['CR', 'DZ'],
    'Room-10': ['CS', 'EA']
}


def do_task3():
    rooms = output_bins['365day']['rooms'] = OrderedDict()
    rooms_s = output_bins['summer']['rooms'] = OrderedDict()
    rooms_w = output_bins['winter']['rooms'] = OrderedDict()

    for name, info in Room_info.items():
        col = info[0]
        # Sum across 365day in 8 time-blocks
        sum_365day = rooms[name] = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        for i in xrange(0, 365):
            time_blocks = data_bins['24hr']['day_'+str(i+1)]['3hr']
            l1 = []
            for block in time_blocks:
                print (block[0])
                l1.append(map(lambda x: float(x[col_map()]), block))

            reduced_blocks = map(lambda x: sum(x), l1)
            sum_365day = [sum(x) for x in zip(sum_365day, reduced_blocks)]

        """
        for i in xrange(0, 365):
            time_blocks = data_bins['24hr']['day_'+str(i+1)]['3hr']
            reduced_blocks = map(lambda x: sum(x), time_blocks)
            for j in xrange(0, 248, 8):
                for k in xrange(0, 8):
                    sum_365day[k] = reduced_blocks[j] + reduced_blocks[j+8]

            #for block in reduced_blocks:
            #    sum_365day[num] += sum(map(lambda x:float( x[col_map[col]] ), block))
        """
        # Sum across the summer and winter in 8 time-blocks
        sum_summer = rooms_s[name] = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        sum_winter = rooms_w[name] = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

        # Generate 3hr bins.
        _3hr_summer_bins = []
        _3hr_winter_bins = []
        summer_data = data_bins['summer']
        winter_data = data_bins['winter']
        for i in xrange(0, len(summer_data), 3):
            _3hr_summer_bins.append(summer_data[i:i+3])

        tmp_s = []
        for i in xrange(0, len(_3hr_summer_bins), 8):
            tmp_s.append(_3hr_summer_bins[i:i+8])
        _3hr_summer_bins = tmp_s

        for j in xrange(0, len(winter_data), 3):
            _3hr_winter_bins.append(winter_data[j:j+3])

        tmp_w = []
        for i in xrange(0, len(_3hr_winter_bins), 8):
            tmp_w.append(_3hr_winter_bins[i:i+8])
        _3hr_winter_bins = tmp_w

        #print len(_3hr_summer_bins)
        #print len(_3hr_winter_bins)

        for i in xrange(0, 744, 24):
            #time_blocks_s = data_bins['summer']['day_'+str(i)]['3hr'] #check
            #time_blocks_w = data_bins['winter']['day_'+str(i)]['3hr'] #check
            #print len(_3hr_summer_bins[0])
            #print len(_3hr_winter_bins)
            #for num, (block_s, block_w) in enumerate(zip(_3hr_summer_bins, _3hr_winter_bins)):
            for i in xrange(0, 248, 24):
                for j in xrange(0, 8):
                    sum_summer[num] += sum(map(lambda x:float( x[col_map[col]]
                        ), block_s[0]))
                    sum_winter[num] += sum(map(lambda x:float( x[col_map[col]]), block_w))

    return output_bins


def do_task4():
    output_bins
    for i in xrange(0, 365):
        days = output_bins['365day']['days']
        rooms = days['day_'+str(i+1)]['rooms'] = OrderedDict()
        for name, info in Room_info.items():
            room = rooms[name] = []
            col = info[1]
            rows_list = data_bins['24hr']['day_'+str(i+1)]['1hr']
            room.append(get_max(col, rows_list))
            room.append(get_min(col, rows_list))
            room.append(get_mean(col, rows_list))

            # Calculating 'Damping Factor'(df)
            inside_temp_range = get_range(col, rows_list)
            #print output_bins['365day']['days']['day_1']['rooms'].keys()
            outside_temp_range = output_bins['365day']['days']['day_'+str(i+1)]['B'][3]
            df = ((outside_temp_range - inside_temp_range)/outside_temp_range)*100

            room.append(inside_temp_range)
            room.append(df)
            #print room
    return output_bins

def prepare_task4():
    sheet = "Task-4"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.create_sheet()
    ws.title = sheet
    col_names = ['Day'] + ['Maximum', 'Minimum', 'Mean', 'Range'] * 10
    ws.append(col_names)
    ws.freeze_panes = 'A2'
    wb.save("Output.xlsx")

def export_task4(filename):
    sheet = "Task-4"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.get_sheet_by_name(sheet)

    for i in xrange(0, 365):
        _data = [filename+"_day"+str(i+1)]
        for key, value in output_bins['365day']['days']['day_'+str(i+1)]['rooms'].items():
            _data.extend([value[0], value[1], value[2], value[3]])
        ws.append(_data)
    wb.save("Output.xlsx")


def do_task5():
    corr = output_bins['correlations'] = {}
    corr['rooms'] = OrderedDict()
    tmp = {}
    for name, info in Room_info.items():
        r2_max = corr_util('B', 'max', name, 'max')
        r2_min = corr_util('B', 'min', name, 'min')
        r2_mean = corr_util('B', 'mean', name, 'mean')
        output_bins['correlations']['rooms'][name] = [[r2_max, r2_min, r2_mean]]
        #print room
    return output_bins

def do_task6():
    data = data_bins['365day']
    tmp1 = []
    for i in xrange(0, 8759):
        item = data[i]
        _val = item[col_map['EG']]
        if _val is not '':
            tmp1.append(float(_val))

    for room_name in Room_info.keys():
        mean_temps = []
        for i in xrange(0, 365):
            mean_temps.append(output_bins['365day']['days']['day_'+str(i+1)]['rooms'][room_name][2])

        corr = get_correlation(tmp1, mean_temps)
        output_bins['correlations']['rooms'][room_name].append(corr)
    return output_bins

def prepare_task5_6():
    sheet = "Task-5_6"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.create_sheet()
    ws.title = sheet
    col_names = ['File'] + ['Maximum', 'Minimum', 'Mean', 'EG'] * 10
    ws.append(col_names)
    ws.freeze_panes = 'A2'
    wb.save("Output.xlsx")

def export_task5_6(filename):
    sheet = "Task-5_6"
    wb = openpyxl.load_workbook("Output.xlsx")
    ws = wb.get_sheet_by_name(sheet)

    _data = [filename]
    for value in output_bins['correlations']['rooms'].values():
        _data.extend(value[0])
        _data.append(value[1])
    ws.append(_data)
    wb.save("Output.xlsx")


# Utility Definitions
def corr_util(col, type_1, room, type_2):
    _indexes = {'max': 0, 'min': 1, 'mean': 2, 'range': 3}
    values_x = []
    values_y = []
    for i in xrange(0, 364):
        x = output_bins['365day']['days']['day_'+str(i+1)][col][_indexes[type_1]]
        y = output_bins['365day']['days']['day_'+str(i+1)]['rooms'][room][_indexes[type_2]]
        values_x.append(float(x))
        values_y.append(float(y))
    return get_correlation(values_x, values_y)

def get_correlation(values_x, values_y):
    _mean_x = sum(values_x)/float(len(values_x))
    _mean_y = sum(values_y)/float(len(values_y))
    n = len(values_x) # equals to len(values_y) also

    _sum1 = 0.0
    _sum2_x = 0.0
    _sum2_y = 0.0
    for j in xrange(0, n):
        diff_x = values_x[j] - _mean_x
        diff_y = values_y[j] - _mean_y
        _sum1 += (diff_x * diff_y)
        _sum2_x += math.pow(diff_x, 2)
        _sum2_y += math.pow(diff_y, 2)

    _covariance = _sum1/float((n-1))
    SD_x = math.sqrt(_sum2_x/float((n-1)))
    SD_y = math.sqrt(_sum2_y/float((n-1)))
    # TODO: Handle negative square roots here
    _correlation = _covariance/(SD_x * SD_y)
    return math.pow(_correlation, 2)

def get_max(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return max(tmp)

def get_min(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return min(tmp)

def get_mean(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return sum(tmp)/float(len(tmp))

def get_range(col_letter, rows_list):
    return get_max(col_letter, rows_list) - get_min(col_letter, rows_list)

#def get_damping_factor(col_letter, rows_list):
#    return get_range()





def check_output():
    #print colored("\n\n---------- Re-Checking the Output ----------", "green")
    #print "\n" + str(len(output_bins.keys()))
    #print "\n" + str(output_bins['365day']['days'].keys())
    #print "\n" + str(output_bins['365day']['days']['day_1'].keys())
    #print "\n" + str(output_bins['365day']['days']['day_1']['rooms'].items())
    #print "\n" + str(output_bins['365day']['days']['day_1']['B'])
    #print "\n" + str(output_bins['365day']['days']['day_1']['rooms']['Room-1'])
    print "\n" + str(output_bins['summer']['days']['day_1']['walls'].items())
    print "\n" + str(output_bins['winter']['days']['day_1']['walls'].items())
    #print "\n" + str(output_bins['365day']['days']['day_1']['walls']['Wall-1'])
    #print "\n" + str(output_bins['correlations']['rooms']['Room-1'])
    #print "\n" + str(len(output_bins['correlations']['rooms'].items()))
    #print "\n" + str(output_bins['correlations']['rooms'].items())

    #print output_bins['winter'].keys()
    #print len(output_bins['winter'])
    #print output_bins['winter']['wall_1'].keys()
    #print len(output_bins['winter']['wall_1'])
    #print output_bins['winter']['wall_1']['H'].keys()
    #print len(output_bins['winter']['wall_1']['H'])
    #print output_bins['summer']['wall_2']['S']['day_1']
    #print output_bins['summer']['wall_2']['T']['day_1']

    #print colored("\n---------- Terminating ---------- \n\n", "green")


def get_filepaths(_dir):
    tmp = OrderedDict()
    for dirpath, _, filenames in os.walk(_dir):
        for f in filenames:
            if f.endswith('.csv'):
                filename = os.path.splitext(f)[0]
                tmp[int(filename)] = os.path.abspath(os.path.join(dirpath, f))
    tmp = map(lambda x: tmp[x], sorted(tmp))  # Sort by filename
    return tmp

# Driver Programme
if __name__ == "__main__":

    files_no = input('Number of input files to process: ')
    input_dir = '/home/gen/Documents/jagan/batch-4/input/'
    output_dir = '/home/gen/Documents/jagan/batch-4/output/'

    os.chdir(output_dir)

    filepaths = get_filepaths(input_dir)
    filepaths = filepaths[:files_no]
    #filepaths = ['/home/jagannadh/Desktop/project-data/input/2.csv',
    #                '/home/jagannadh/Desktop/project-data/input/100.csv']

    # Initilization Methods
    create_excel_workbook()
    prepare_task1()
    prepare_task2()
    prepare_task4()
    prepare_task5_6()

    for filepath in filepaths:
        complete_filename = os.path.basename(filepath)
        filename = os.path.splitext(complete_filename)[0]

        print ">> Processing " +  complete_filename

        #os.chdir(input_path)

        # Global Variables
        col_map = {}
        col_name_map = {}
        col_names = []
        data_bins = {}
        output_bins = {}

        #print filepath
        #print data_bins, output_bins

        # Methods
        data = read_csvfile(filepath)
        create_data_bins(data)
        create_data_groups()

        #recheck_data()
        #check_output()

        do_task1()
        do_task2()
        #do_task3()
        do_task4()
        do_task5()
        do_task6()
        export_task1(filename)
        export_task2(filename)
        export_task4(filename)
        export_task5_6(filename)
